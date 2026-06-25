"""
Import ILUP for Drylands Project Excel data into public/drylands-data.json
Merges with any existing manually-added records (those with _id < 100000 or string IDs).
"""
import openpyxl
import json
import os
from datetime import datetime

script_dir = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(script_dir, "ILUP_for_Drylands_Project_-_all_versions_-_labels_-_2026-05-12-08-42-15.xlsx")
OUTPUT_FILE = os.path.join(script_dir, "public", "drylands-data.json")

# District lookup by ward number (from field knowledge)
WARD_DISTRICT_MAP = {
    "4": "Chivi", "5": "Chivi", "6": "Chivi", "7": "Chivi", "8": "Chivi",
    "11": "Gutu", "12": "Gutu", "13": "Gutu", "14": "Gutu",
    "17": "Chipinge", "18": "Chipinge", "19": "Chipinge", "20": "Chipinge",
    "1": "Bikita", "2": "Bikita", "3": "Bikita",
    "9": "Zaka", "10": "Zaka",
    "15": "Buhera", "16": "Buhera",
}

def infer_district(ward_name, village_location=""):
    w = str(ward_name or "").strip()
    v = str(village_location or "").lower()
    # Extract ward number
    for token in w.split():
        clean = token.replace("Ward", "").replace("ward", "").strip()
        if clean.isdigit() and clean in WARD_DISTRICT_MAP:
            return WARD_DISTRICT_MAP[clean]
    if clean.isdigit():
        return WARD_DISTRICT_MAP.get(clean, "Chivi")
    # Fallback by village keywords
    if "musikavanhu" in v or "chipinge" in v: return "Chipinge"
    if "gutu" in v: return "Gutu"
    if "chivi" in v: return "Chivi"
    if "bikita" in v: return "Bikita"
    if "zaka" in v: return "Zaka"
    if "buhera" in v: return "Buhera"
    return "Chivi"

def parse_date(val):
    if val is None: return ""
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"]:
        try:
            return datetime.strptime(s[:10], fmt[:len(fmt.split()[0])]).strftime("%Y-%m-%d")
        except: pass
    return s[:10]

def parse_datetime(val):
    if val is None: return ""
    if isinstance(val, datetime): return val.strftime("%Y-%m-%dT%H:%M:%S")
    return str(val).replace(" ", "T")[:19]

def cell(ws, row, col):
    v = ws.cell(row, col).value
    return str(v).strip() if v is not None else ""

def import_ilup_records(ws):
    records = []
    for row in range(2, ws.max_row + 1):
        _id = ws.cell(row, 131).value
        if not _id:
            continue
        _id = int(_id)

        ward_name_raw = cell(ws, row, 5)
        ward_display = f"Ward {ward_name_raw}" if ward_name_raw.isdigit() else ward_name_raw
        village = cell(ws, row, 6)
        district = infer_district(ward_name_raw, village)

        # Recommended interventions — may be space/comma separated
        interventions_raw = cell(ws, row, 111)
        interventions = [i.strip() for i in interventions_raw.replace("  ", " ").split() if i.strip()] if interventions_raw else []

        # Photo URLs
        photos = []
        for photo_url_col in [125, 127, 129]:
            url = cell(ws, row, photo_url_col)
            if url and url.startswith("http"):
                photos.append(url)

        record = {
            "_id": _id,
            "_uuid": cell(ws, row, 132),
            "_submission_time": parse_datetime(ws.cell(row, 133).value),
            "date_of_observation": parse_date(ws.cell(row, 3).value),
            "enumerator_name": cell(ws, row, 4),
            "ward_name": ward_display,
            "village_location": village,
            "coordinates": cell(ws, row, 7),
            "dist": district,
            "district": district,
            # Land / environment
            "area_type": cell(ws, row, 12),
            "dominant_soil_type": cell(ws, row, 21),
            "distance_to_road_m": cell(ws, row, 25),
            "vegetation_condition": cell(ws, row, 26).strip(),
            "estimated_vegetation_cover": ws.cell(row, 28).value,
            "invasive_species_present": cell(ws, row, 29),
            "invasive_species": cell(ws, row, 30),
            "current_land_cover": cell(ws, row, 31),
            # Erosion
            "soil_erosion_present": cell(ws, row, 39),
            "erosion_type": cell(ws, row, 40),
            "erosion_severity": cell(ws, row, 45),
            "erosion_expanding": cell(ws, row, 47),
            "assets_threatened": cell(ws, row, 48),
            # Water
            "water_sources": cell(ws, row, 56),
            "water_quality": cell(ws, row, 64),
            "siltation_evidence": cell(ws, row, 69),
            "wetland_cultivation": cell(ws, row, 70),
            # Climate
            "climate_indicators": cell(ws, row, 72),
            "flood_prone": cell(ws, row, 80),
            "fire_evidence": cell(ws, row, 82),
            "drought_stress": cell(ws, row, 84),
            "drought_notes": cell(ws, row, 85),
            # Livelihoods
            "livelihood_activities": cell(ws, row, 86),
            "grazing_pressure": cell(ws, row, 95),
            "land_use_conflicts": cell(ws, row, 100),
            "ecologically_compatible": cell(ws, row, 107),
            "ecologically_compatible_notes": cell(ws, row, 108),
            # Priority & interventions
            "priority_level": cell(ws, row, 109),
            "priority_notes": cell(ws, row, 110),
            "recommended_interventions": interventions,
            "intervention_cost_category": cell(ws, row, 122),
            "intervention_cost_notes": cell(ws, row, 123),
            "notes": cell(ws, row, 130),
            # Photos
            "photo_urls": photos,
            # Source marker
            "_source": "ilup_excel",
        }
        records.append(record)
        print(f"  Row {row}: _id={_id}, ward={ward_display}, village={village}, district={district}, date={record['date_of_observation']}")

    return records

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        return

    print(f"Loading: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb.active
    print(f"Found {ws.max_row - 1} data rows in sheet '{ws.title}'")

    # Parse ILUP records
    ilup_records = import_ilup_records(ws)
    print(f"\nParsed {len(ilup_records)} ILUP records.")

    # Load existing drylands-data.json
    existing_records = []
    if os.path.exists(OUTPUT_FILE):
        with open(OUTPUT_FILE, "r", encoding="utf-8") as f:
            existing_data = json.load(f)
            existing_records = existing_data.get("records", [])
        print(f"Existing records in drylands-data.json: {len(existing_records)}")
    
    # Merge: keep manually-added records (those NOT from ILUP), add ILUP records
    ilup_ids = {str(r["_id"]) for r in ilup_records}
    manual_records = [r for r in existing_records if str(r.get("_id", "")) not in ilup_ids and r.get("_source") != "ilup_excel"]
    print(f"Preserving {len(manual_records)} manually-added records.")

    combined = ilup_records + manual_records
    combined.sort(key=lambda r: str(r.get("_submission_time", "")), reverse=True)

    output = {"count": len(combined), "records": combined}
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nOK: Saved {len(combined)} drylands records to {OUTPUT_FILE}")
    print(f"  - {len(ilup_records)} from ILUP Excel")
    print(f"  - {len(manual_records)} manually added")

if __name__ == "__main__":
    main()
